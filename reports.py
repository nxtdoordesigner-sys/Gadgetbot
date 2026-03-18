import os
import tempfile
from datetime import datetime, timezone
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter
from supabase_client import supabase

# ── Colors ────────────────────────────────────────────────
DARK = "1A1A2E"
ACCENT = "E94560"
LIGHT_GRAY = "F5F5F5"
WHITE = "FFFFFF"
GREEN = "27AE60"
ORANGE = "E67E22"
RED = "E74C3C"


def style_header(cell, bg=DARK, fg=WHITE, bold=True, size=11):
    cell.font = Font(bold=bold, color=fg, size=size, name="Arial")
    cell.fill = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def style_cell(cell, bold=False, color="000000", bg=None, align="left", size=10):
    cell.font = Font(bold=bold, color=color, size=size, name="Arial")
    if bg:
        cell.fill = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(horizontal=align, vertical="center")


def thin_border():
    thin = Side(style="thin", color="DDDDDD")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def apply_border(ws, min_row, max_row, min_col, max_col):
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            cell.border = thin_border()


def set_col_widths(ws, widths: dict):
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def add_title_block(ws, title: str, subtitle: str, row=1):
    ws.merge_cells(f"A{row}:H{row}")
    cell = ws[f"A{row}"]
    cell.value = title
    cell.font = Font(bold=True, color=WHITE, size=16, name="Arial")
    cell.fill = PatternFill("solid", start_color=DARK)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 35

    ws.merge_cells(f"A{row+1}:H{row+1}")
    sub = ws[f"A{row+1}"]
    sub.value = subtitle
    sub.font = Font(color="888888", size=10, name="Arial")
    sub.fill = PatternFill("solid", start_color=LIGHT_GRAY)
    sub.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row+1].height = 20


# ── 1. ORDERS REPORT ──────────────────────────────────────
def build_orders_sheet(wb: Workbook):
    ws = wb.create_sheet("📦 Orders")
    now = datetime.now(timezone.utc)

    res = supabase.table("orders").select("*").order("created_at", desc=True).execute()
    orders = res.data or []

    add_title_block(ws, "VoltStore — Orders Report", f"Generated: {now.strftime('%d %b %Y, %H:%M')} UTC")

    headers = ["#", "Order ID", "Customer", "Phone", "Delivery Address", "Items", "Total (₦)", "Status", "Date"]
    header_row = 4
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col, value=h)
        style_header(cell, bg=ACCENT)
    ws.row_dimensions[header_row].height = 22

    status_colors = {"pending": ORANGE, "confirmed": GREEN, "delivered": "2980B9", "cancelled": RED}

    for i, order in enumerate(orders):
        row = header_row + 1 + i
        bg = WHITE if i % 2 == 0 else LIGHT_GRAY
        items_text = ", ".join([f"{it['title']} x{it['quantity']}" for it in order.get("items", [])])

        data = [
            i + 1,
            order["id"],
            order.get("customer_name", ""),
            order.get("phone_number", "N/A"),
            order.get("location", "N/A"),
            items_text,
            order.get("total", 0),
            order.get("status", "").capitalize(),
            order.get("created_at", "")[:10],
        ]
        for col, val in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=val)
            style_cell(cell, bg=bg)
            if col == 7:  # Total
                cell.number_format = '#,##0'
                cell.alignment = Alignment(horizontal="right")
            if col == 8:  # Status
                color = status_colors.get(order.get("status", ""), "000000")
                cell.font = Font(bold=True, color=color, size=10, name="Arial")
                cell.alignment = Alignment(horizontal="center")

    apply_border(ws, header_row, header_row + len(orders), 1, len(headers))

    # Summary block
    summary_row = header_row + len(orders) + 2
    ws.cell(row=summary_row, column=1, value="SUMMARY").font = Font(bold=True, size=11, name="Arial")
    totals = [
        ("Total Orders", len(orders)),
        ("Pending", len([o for o in orders if o["status"] == "pending"])),
        ("Confirmed", len([o for o in orders if o["status"] == "confirmed"])),
        ("Delivered", len([o for o in orders if o["status"] == "delivered"])),
        ("Cancelled", len([o for o in orders if o["status"] == "cancelled"])),
        ("Total Revenue (₦)", sum(o["total"] for o in orders if o["status"] in ["confirmed", "delivered"])),
    ]
    for j, (label, val) in enumerate(totals):
        r = summary_row + 1 + j
        ws.cell(row=r, column=1, value=label).font = Font(name="Arial", size=10)
        cell = ws.cell(row=r, column=2, value=val)
        cell.font = Font(bold=True, name="Arial", size=10)
        if "Revenue" in label:
            cell.number_format = '#,##0'

    set_col_widths(ws, {"A": 5, "B": 10, "C": 20, "D": 15, "E": 30, "F": 40, "G": 15, "H": 12, "I": 12})


# ── 2. INVENTORY SHEET ────────────────────────────────────
def build_inventory_sheet(wb: Workbook):
    ws = wb.create_sheet("🗂 Inventory")
    now = datetime.now(timezone.utc)

    res = supabase.table("books").select("*").order("category").execute()
    products = res.data or []

    add_title_block(ws, "VoltStore — Inventory", f"Generated: {now.strftime('%d %b %Y, %H:%M')} UTC")

    headers = ["ID", "Product", "Brand", "Category", "Condition", "Price (₦)", "Stock", "Negotiable", "Base Price (₦)", "Status"]
    header_row = 4
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col, value=h)
        style_header(cell, bg=DARK)
    ws.row_dimensions[header_row].height = 22

    for i, p in enumerate(products):
        row = header_row + 1 + i
        bg = WHITE if i % 2 == 0 else LIGHT_GRAY
        stock = p.get("stock_qty", 0)
        in_stock = p.get("in_stock", False)

        data = [
            p["id"], p["title"], p.get("author", ""),
            p.get("category", ""), p.get("condition", "Brand New"),
            p.get("price", 0), stock,
            "Yes" if p.get("negotiable") else "No",
            p.get("base_price", 0) or 0,
            "✅ In Stock" if in_stock else "❌ Out of Stock",
        ]
        for col, val in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=val)
            style_cell(cell, bg=bg)
            if col in [6, 9]:
                cell.number_format = '#,##0'
                cell.alignment = Alignment(horizontal="right")
            if col == 7:  # Stock qty
                cell.alignment = Alignment(horizontal="center")
                if stock <= 2 and in_stock:
                    cell.font = Font(bold=True, color=RED, name="Arial", size=10)
            if col == 10:  # Status
                color = GREEN if in_stock else RED
                cell.font = Font(bold=True, color=color, name="Arial", size=10)
                cell.alignment = Alignment(horizontal="center")

    apply_border(ws, header_row, header_row + len(products), 1, len(headers))
    set_col_widths(ws, {"A": 6, "B": 30, "C": 15, "D": 15, "E": 15, "F": 15, "G": 8, "H": 12, "I": 15, "J": 14})


# ── 3. REVENUE REPORT ─────────────────────────────────────
def build_revenue_sheet(wb: Workbook):
    ws = wb.create_sheet("💰 Revenue")
    now = datetime.now(timezone.utc)

    res = supabase.table("orders").select("*").execute()
    orders = [o for o in (res.data or []) if o["status"] in ["confirmed", "delivered"]]

    add_title_block(ws, "VoltStore — Revenue Report", f"Generated: {now.strftime('%d %b %Y, %H:%M')} UTC")

    # Monthly breakdown
    monthly = {}
    for o in orders:
        month = o["created_at"][:7]
        monthly[month] = monthly.get(month, 0) + o["total"]

    ws.cell(row=4, column=1, value="Monthly Revenue").font = Font(bold=True, size=12, name="Arial")

    headers = ["Month", "Revenue (₦)", "Orders"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col, value=h)
        style_header(cell, bg=ACCENT)

    monthly_orders = {}
    for o in orders:
        month = o["created_at"][:7]
        monthly_orders[month] = monthly_orders.get(month, 0) + 1

    for i, (month, rev) in enumerate(sorted(monthly.items(), reverse=True)):
        row = 6 + i
        bg = WHITE if i % 2 == 0 else LIGHT_GRAY
        ws.cell(row=row, column=1, value=month).font = Font(name="Arial", size=10)
        cell = ws.cell(row=row, column=2, value=rev)
        cell.number_format = '#,##0'
        style_cell(cell, bg=bg)
        ws.cell(row=row, column=3, value=monthly_orders.get(month, 0)).font = Font(name="Arial", size=10)

    # Totals
    total_row = 6 + len(monthly) + 1
    ws.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True, name="Arial", size=11)
    total_cell = ws.cell(row=total_row, column=2, value=f"=SUM(B6:B{total_row-1})")
    total_cell.number_format = '#,##0'
    total_cell.font = Font(bold=True, color=ACCENT, name="Arial", size=11)

    # Category breakdown
    cat_row = total_row + 3
    ws.cell(row=cat_row, column=1, value="Revenue by Category").font = Font(bold=True, size=12, name="Arial")

    cat_headers = ["Category", "Revenue (₦)", "Units Sold"]
    for col, h in enumerate(cat_headers, 1):
        cell = ws.cell(row=cat_row + 1, column=col, value=h)
        style_header(cell, bg=DARK)

    cat_revenue = {}
    cat_units = {}
    for o in orders:
        for item in o.get("items", []):
            cat = item.get("category", "Other")
            cat_revenue[cat] = cat_revenue.get(cat, 0) + item["price"] * item["quantity"]
            cat_units[cat] = cat_units.get(cat, 0) + item["quantity"]

    for i, (cat, rev) in enumerate(sorted(cat_revenue.items(), key=lambda x: x[1], reverse=True)):
        row = cat_row + 2 + i
        bg = WHITE if i % 2 == 0 else LIGHT_GRAY
        ws.cell(row=row, column=1, value=cat).font = Font(name="Arial", size=10)
        cell = ws.cell(row=row, column=2, value=rev)
        cell.number_format = '#,##0'
        style_cell(cell, bg=bg)
        ws.cell(row=row, column=3, value=cat_units.get(cat, 0)).font = Font(name="Arial", size=10)

    apply_border(ws, 5, 5 + len(monthly) + 1, 1, 3)
    set_col_widths(ws, {"A": 20, "B": 20, "C": 15})


# ── 4. CUSTOMER LIST ──────────────────────────────────────
def build_customers_sheet(wb: Workbook):
    ws = wb.create_sheet("👥 Customers")
    now = datetime.now(timezone.utc)

    res = supabase.table("orders").select("*").execute()
    all_orders = res.data or []

    # Aggregate by customer
    customers = {}
    for o in all_orders:
        tid = o["telegram_id"]
        if tid not in customers:
            customers[tid] = {
                "name": o["customer_name"],
                "phone": o.get("phone_number", "N/A"),
                "telegram_id": tid,
                "orders": 0,
                "spent": 0,
                "last_order": o["created_at"][:10],
            }
        customers[tid]["orders"] += 1
        if o["status"] in ["confirmed", "delivered"]:
            customers[tid]["spent"] += o["total"]

    add_title_block(ws, "VoltStore — Customer List", f"Generated: {now.strftime('%d %b %Y, %H:%M')} UTC")

    headers = ["#", "Name", "Phone", "Telegram ID", "Total Orders", "Total Spent (₦)", "Last Order"]
    header_row = 4
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col, value=h)
        style_header(cell, bg=DARK)

    sorted_customers = sorted(customers.values(), key=lambda x: x["spent"], reverse=True)
    for i, c in enumerate(sorted_customers):
        row = header_row + 1 + i
        bg = WHITE if i % 2 == 0 else LIGHT_GRAY
        data = [i+1, c["name"], c["phone"], c["telegram_id"], c["orders"], c["spent"], c["last_order"]]
        for col, val in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=val)
            style_cell(cell, bg=bg)
            if col == 6:
                cell.number_format = '#,##0'
                cell.alignment = Alignment(horizontal="right")

    apply_border(ws, header_row, header_row + len(customers), 1, len(headers))
    set_col_widths(ws, {"A": 5, "B": 25, "C": 18, "D": 18, "E": 14, "F": 18, "G": 14})


# ── 5. LOW STOCK ALERT ────────────────────────────────────
def build_lowstock_sheet(wb: Workbook):
    ws = wb.create_sheet("⚠️ Low Stock")
    now = datetime.now(timezone.utc)

    res = supabase.table("books").select("*").lte("stock_qty", 3).eq("in_stock", True).execute()
    products = res.data or []

    add_title_block(ws, "VoltStore — Low Stock Alert", f"Generated: {now.strftime('%d %b %Y, %H:%M')} UTC")

    headers = ["ID", "Product", "Brand", "Category", "Stock Left", "Price (₦)"]
    header_row = 4
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col, value=h)
        style_header(cell, bg=RED)

    for i, p in enumerate(products):
        row = header_row + 1 + i
        stock = p.get("stock_qty", 0)
        bg = "FFEBEB" if stock <= 1 else "FFF3CD"
        data = [p["id"], p["title"], p.get("author", ""), p.get("category", ""), stock, p.get("price", 0)]
        for col, val in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=val)
            style_cell(cell, bg=bg)
            if col == 5:
                cell.font = Font(bold=True, color=RED if stock <= 1 else ORANGE, name="Arial", size=10)
                cell.alignment = Alignment(horizontal="center")
            if col == 6:
                cell.number_format = '#,##0'

    if not products:
        ws.cell(row=5, column=1, value="✅ All products are well stocked!").font = Font(color=GREEN, bold=True, name="Arial")

    apply_border(ws, header_row, header_row + max(len(products), 1), 1, len(headers))
    set_col_widths(ws, {"A": 6, "B": 35, "C": 18, "D": 15, "E": 12, "F": 15})


# ── MAIN EXPORT FUNCTION ──────────────────────────────────
def generate_report(report_type: str = "full") -> str:
    """Generate Excel report and return file path."""
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    now = datetime.now(timezone.utc)
    fname = f"VoltStore_Report_{now.strftime('%Y%m%d_%H%M')}.xlsx"
    fpath = os.path.join(tempfile.gettempdir(), fname)

    if report_type in ["full", "orders"]:
        build_orders_sheet(wb)
    if report_type in ["full", "inventory"]:
        build_inventory_sheet(wb)
    if report_type in ["full", "revenue"]:
        build_revenue_sheet(wb)
    if report_type in ["full", "customers"]:
        build_customers_sheet(wb)
    if report_type in ["full", "lowstock"]:
        build_lowstock_sheet(wb)

    wb.save(fpath)
    return fpath
